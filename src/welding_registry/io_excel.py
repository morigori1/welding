from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Literal

import pandas as pd

from .field_map import get_header_map, DATE_COLUMNS

# NOTE: This module handles two distinct layout families:
# 1) Standard headered tables (read_sheet/to_canonical)
# 2) "Vertical blocks" sheets where a person's name/登録番号 spans rows and
#    multiple category blocks (JIS/BOILER) live in column ranges.
#
# This patch adds lightweight auto-detection for the vertical layout so that
# diverse patterns can be ingested without hardcoding column letters.


@dataclass
class SheetSummary:
    name: str
    n_rows: int
    n_cols: int
    headers: List[str]


def _engine_for(path: Path) -> Literal["openpyxl", "xlrd"]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return "openpyxl"
    return "xlrd"


def list_sheets(xls_path: Path) -> List[str]:
    with pd.ExcelFile(xls_path, engine=_engine_for(xls_path)) as xf:
        return list(map(str, xf.sheet_names))


def _detect_header_row(df: pd.DataFrame) -> Optional[int]:
    """Heuristically find the header row by matching known Japanese header tokens.
    Returns a 0-based row index within the originally read frame, or None.
    """
    header_tokens = set(get_header_map().keys())
    for i in range(min(20, len(df))):
        values = set(str(v).strip() for v in df.iloc[i].tolist())
        # Count meaningful header tokens present on this row
        hit = sum(1 for v in values if v in header_tokens)
        if hit >= 2:  # at least two known headers on same row
            return i
    return None


def read_sheet(
    xls_path: Path, sheet_name: str | int, header_row_override: int | None = None
) -> Tuple[pd.DataFrame, Optional[int]]:
    # Read a sample with no header to detect the header row
    df_probe = pd.read_excel(
        xls_path, sheet_name=sheet_name, header=None, engine=_engine_for(xls_path)
    )  # type: ignore[call-overload]
    header_row: Optional[int]
    if header_row_override is not None:
        header_row = header_row_override
    else:
        header_row = _detect_header_row(df_probe)
        if header_row is None:
            # Fallback: first non-empty row; if none, use 0 to keep reading
            counts = df_probe.notna().sum(axis=1)
            nz = counts[counts > 0]
            header_row = int(nz.index.min()) if not nz.empty else 0

    df = pd.read_excel(  # type: ignore[call-overload]
        xls_path,
        sheet_name=sheet_name,
        header=header_row,
        dtype="object",
        engine=_engine_for(xls_path),
    )
    # Drop completely empty columns/rows
    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    # Normalize column labels to strings without surrounding spaces
    df.columns = [str(c).strip() for c in df.columns]
    # Demote header cells that look like dates to Unnamed: <idx> (avoid accidental mapping/"ずれ")
    import re as _re

    new_cols: list[str] = []
    for i, c in enumerate(df.columns):
        try:
            ts = pd.to_datetime(c, errors="coerce")
            if pd.notna(ts) and _re.match(r"^\d{2,4}([/.-]\d{1,2}){1,2}", str(c)):
                new_cols.append(f"Unnamed: {i}")
                continue
        except Exception:
            pass
        new_cols.append(str(c))
    df.columns = new_cols
    return df, header_row


def summarize(xls_path: Path) -> List[SheetSummary]:
    summaries: List[SheetSummary] = []
    for s in list_sheets(xls_path):
        df, header_row = read_sheet(xls_path, s)
        summaries.append(
            SheetSummary(
                name=str(s), n_rows=len(df), n_cols=df.shape[1], headers=list(map(str, df.columns))
            )
        )
    return summaries


def to_canonical(df: pd.DataFrame) -> pd.DataFrame:
    """Map Japanese headers to canonical English snake_case names.
    Unmapped columns are preserved with their original labels.
    """
    header_map = get_header_map()

    def _norm(s: str) -> str:
        from .field_map import _norm_token  # type: ignore

        try:
            return _norm_token(s)
        except Exception:
            return str(s).strip()

    mapped_cols = {}
    import re

    for col in df.columns:
        raw = str(col).strip()
        norm = _norm(raw)
        key = header_map.get(raw) or header_map.get(norm)
        if not key:
            # Heuristics for common patterns
            if re.search(r"氏名|名前", norm):
                key = "name"
            elif re.search(r"資格", norm):
                key = "qualification"
            elif re.search(r"番号|No\.?", norm, re.IGNORECASE):
                key = "license_no"
            elif "試験" in norm:
                key = "test_date"
            elif ("取得" in norm and "日" in norm) or ("登録" in norm and "日" in norm):
                key = "first_issue_date"
            elif "継続" in norm and "日" in norm:
                key = "issue_date"
            elif "交付" in norm and "日" in norm:
                key = "issue_date"
            elif "有効" in norm and ("年月日" in norm or "期限" in norm or "満了" in norm):
                key = "expiry_date"
            elif ("交付" in norm) or ("発行" in norm):
                key = "issue_date"
            elif ("有効" in norm) and ("期限" in norm or "満了" in norm):
                key = "expiry_date"
            elif "生年月日" in norm:
                key = "birth_date"
            elif "西暦" in norm or "生年" in norm:
                key = "birth_year_west"
        mapped_cols[col] = key if key else col
    out = df.rename(columns=mapped_cols)

    # Parse dates where possible (avoid struct-like assembly errors by coercing series)
    from .dates_jp import parse_jp_date  # lazy import
    import re as _re

    def _extract_paren_year(val: Any) -> Optional[int]:
        try:
            s = str(val)
        except Exception:
            return None
        m = _re.search(r"\(([^)]*)\)", s)
        if not m:
            return None
        inner = m.group(1)
        m2 = _re.search(r"(\d{2,4})", inner)
        if not m2:
            return None
        yy = m2.group(1)
        try:
            n = int(yy)
        except Exception:
            return None
        if len(yy) <= 2:
            # Assume 2000s for two-digit years (e.g., '(23)' -> 2023)
            base = 2000
            return base + n
        return n

    for c in DATE_COLUMNS:
        if c in out.columns:
            # Coalesce duplicate-named columns first
            idxs = [i for i, col in enumerate(out.columns) if col == c]
            sraw = (
                out.iloc[:, idxs[0]]
                if len(idxs) == 1
                else out.iloc[:, idxs].bfill(axis=1).iloc[:, 0]
            )
            s: pd.Series
            try:
                s = pd.to_datetime(sraw, errors="coerce")
            except Exception:
                # Some mixed-type columns can trigger unit-assembly errors; fall back to per-cell parsing
                s = pd.Series([pd.NaT] * len(out))
            # For remaining NaT, try JP-specific parser and strip memo like '(23)'
            if s.isna().any():

                def _coerce(v):
                    if v is None or (isinstance(v, float) and pd.isna(v)):
                        return None
                    t = str(v).strip()
                    # remove trailing parenthetical notes like '(23)'
                    t = t.split("(")[0].strip()
                    dt = parse_jp_date(t)
                    return pd.to_datetime(dt) if dt else pd.NaT

                s2 = sraw.map(_coerce)
                s = s.combine_first(s2)  # type: ignore[assignment]
            out[c] = s
            # If this is expiry_date, also extract acquisition year from trailing parentheses
            if c == "expiry_date":
                acq_series = sraw.map(_extract_paren_year)
                colname = "issue_year"
                if colname in out.columns:
                    out[colname] = pd.Series(out[colname]).combine_first(acq_series)
                else:
                    out[colname] = acq_series
    return out


def write_xlsx(df: pd.DataFrame, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        df.to_excel(xw, index=False)


def write_csv(df: pd.DataFrame, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_path, index=False, encoding="utf-8-sig")


# --- Vertical block reader for name-spanning rows layouts ---


def _col_letter_to_index(col: str) -> int:
    """Convert Excel column letter(s) (e.g., 'A', 'C', 'AA') to 0-based index."""
    col = str(col).strip().upper()
    acc = 0
    for ch in col:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid column letter: {col}")
        acc = acc * 26 + (ord(ch) - ord("A") + 1)
    return acc - 1


def _parse_col_range(expr: str) -> tuple[int, int]:
    """Parse a range like 'C:H' into (start_idx, end_idx_exclusive)."""
    expr = str(expr).strip()
    if ":" not in expr:
        idx = _col_letter_to_index(expr)
        return idx, idx + 1
    left, right = [p.strip() for p in expr.split(":", 1)]
    a = _col_letter_to_index(left)
    b = _col_letter_to_index(right)
    if b < a:
        a, b = b, a
    return a, b + 1


def _index_to_col_letter(idx: int) -> str:
    """Convert 0-based column index to Excel-style letters (0->A, 25->Z, 26->AA)."""
    if idx < 0:
        raise ValueError("index must be non-negative")
    out = []
    n = idx + 1
    while n > 0:
        n, r = divmod(n - 1, 26)
        out.append(chr(ord("A") + r))
    return "".join(reversed(out))


# --- Auto detection helpers for vertical-block layout ---


def _density(series: pd.Series) -> float:
    try:
        return float(series.notna().sum()) / max(1, len(series))
    except Exception:
        return 0.0


def _looks_like_regno_token(s: Optional[str]) -> bool:
    if s is None:
        return False
    import re as _re

    t = str(s).strip().replace(" ", "")
    # Typical patterns: 12345, 12-3456, SE2500123, UE1100123, ME2300710 etc.
    return bool(_re.fullmatch(r"([A-Z]{1,2}\d{6,}|\d{1,6}(-\d{1,6})?)", t))


def _looks_like_dateish_token(s: Optional[str]) -> bool:
    if not s:
        return False
    import re as _re

    t = str(s)
    return ("年" in t and ("月" in t or "日" in t)) or bool(
        _re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{2}[./]\d{1,2}[./]\d{1,2}", t)
    )


def _is_headerish_name_token(s: Optional[str]) -> bool:
    if not s:
        return True
    t = str(s).strip()
    if t == "":
        return True
    if any(k in t for k in ("氏名", "生年", "生年月日", "備考", "計画", "資格")):
        return True
    if t.startswith("(") and t.endswith(")"):
        return True
    if _looks_like_dateish_token(t):
        return True
    return False


def _looks_like_name_token(s: Optional[str]) -> bool:
    if not s:
        return False
    t = str(s).strip()
    if t == "":
        return False
    # Exclude headerish/dateish tokens handled in existing helper below
    if _is_headerish_name_token(t) or _looks_like_dateish_token(t) or _looks_like_regno_token(t):
        return False
    # Heuristic: short-ish text (<= 20 chars) and contains letters/kanji-kana-like
    return len(t) <= 20


def _label_for_block(df_all: pd.DataFrame, c0: int, c1: int) -> str:
    """Pick a human label for a block by scanning header/body text heuristically."""
    head_vals = " ".join(
        str(v) for v in df_all.iloc[:, c0:c1].astype(str).head(6).values.ravel() if pd.notna(v)
    ).lower()
    body_vals = " ".join(
        str(v) for v in df_all.iloc[:, c0:c1].astype(str).head(50).values.ravel() if pd.notna(v)
    ).lower()
    blob = f"{head_vals} {body_vals}"
    # Boiler-like tokens
    if any(tok in blob for tok in ["boiler", "ﾎﾞｲﾗ", "ボイラ", "ボイラー", "a-3f", "a-3v", "futsu"]):
        return "BOILER"
    # JIS-like tokens
    if (
        ("jis" in blob)
        or ("ｊｉｓ" in blob)
        or ("溶接" in blob)
        or any(
            tok in blob
            for tok in [
                "sc-",
                "cn-",
                "mn-",
                "tn-",
                "n-3",
                "se",
                "ue",
                "me",
            ]
        )
    ):
        return "JIS"
    return f"BLOCK{c0 + 1}"


def detect_vertical_layout_df(
    df_raw: pd.DataFrame,
    *,
    max_probe_rows: int = 10,
    min_block_density: float = 0.1,
) -> tuple[int, int, list[tuple[str, tuple[int, int]]]]:
    """Detect (person_col_idx, regno_col_idx, blocks) from a raw header-less DataFrame.

    - regno column: column with highest ratio of registration-number-like tokens.
    - person column: closest column to the left that contains name-like tokens.
    - blocks: contiguous column ranges to the right with sufficient non-null density.
    """
    df = df_raw.dropna(how="all").reset_index(drop=True)
    probe = df.head(max(1, max_probe_rows))

    # Score each column for registration-number-likeness
    reg_scores: list[float] = []
    for ci in range(probe.shape[1]):
        s = probe.iloc[:, ci]
        score = sum(1 for v in s if _looks_like_regno_token(v)) / max(1, len(s))
        reg_scores.append(score)
    regno_idx = int(pd.Series(reg_scores).idxmax()) if reg_scores else 1

    # Choose person column to the left with name-like density
    name_scores: Dict[int, float] = {}
    for ci in range(max(0, regno_idx)):
        s = probe.iloc[:, ci]
        score = sum(1 for v in s if _looks_like_name_token(v)) / max(1, len(s))
        name_scores[ci] = score
    person_idx = (
        max(name_scores.items(), key=lambda item: item[1])[0]
        if name_scores
        else max(0, regno_idx - 1)
    )

    # Build density vector for potential blocks to the right of reg/reg+1
    start_scan = max(regno_idx + 1, person_idx + 1)
    dens = [(_density(df.iloc[:, ci]), ci) for ci in range(start_scan, df.shape[1])]
    # Identify contiguous regions with density above threshold
    blocks: list[tuple[int, int]] = []
    cur_start = None
    for d, ci in dens:
        if d >= min_block_density:
            if cur_start is None:
                cur_start = ci
        else:
            if cur_start is not None and ci - cur_start >= 1:
                blocks.append((cur_start, ci))  # [start, end)
                cur_start = None
    if cur_start is not None:
        blocks.append((cur_start, df.shape[1]))

    # Assign labels
    labels: list[tuple[str, tuple[int, int]]] = []
    for b0, b1 in blocks:
        if b1 - b0 <= 0:
            continue
        lab = _label_for_block(df, b0, b1)
        labels.append((lab, (b0, b1)))

    # Fallbacks if detection produced nothing reasonable
    if not labels:
        # two coarse blocks following typical layout
        labels = [("JIS", (start_scan + 0, min(start_scan + 6, df.shape[1])))]
        if start_scan + 6 < df.shape[1]:
            labels.append(("BOILER", (start_scan + 6, min(start_scan + 9, df.shape[1]))))

    return person_idx, regno_idx, labels


def detect_vertical_layout(
    xls_path: Path,
    sheet: str | int,
    *,
    max_probe_rows: int = 10,
) -> tuple[int, int, list[tuple[str, tuple[int, int]]]]:
    """Read the sheet with no headers and detect vertical layout components."""
    df_raw = pd.read_excel(
        xls_path, sheet_name=sheet, header=None, engine=_engine_for(xls_path), dtype="object"
    )
    return detect_vertical_layout_df(df_raw, max_probe_rows=max_probe_rows)


# --- Print Area utilities ---


def _bounds_from_a1(a1: str) -> Iterable[tuple[int, int, int, int]]:
    """Yield (r0, r1_excl, c0, c1_excl) from an A1 string (may contain multiple areas)."""
    from openpyxl.utils.cell import range_boundaries  # type: ignore[import-untyped]

    parts = [p.strip() for p in str(a1).split(",") if p.strip()]
    for p in parts:
        if "!" in p:
            p = p.split("!", 1)[1]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(p)
            r0 = int(min_row) - 1
            r1 = int(max_row)  # inclusive -> exclusive
            c0 = int(min_col) - 1
            c1 = int(max_col)
            yield (r0, r1, c0, c1)
        except Exception:
            continue


def get_print_areas(xls_path: Path, sheet_name: str | int) -> list[tuple[int, int, int, int]]:
    """Return list of print areas for a sheet as (r0, r1_excl, c0, c1_excl).

    Supports both .xlsx (openpyxl) and .xls (xlrd). Falls back to empty list if not found.
    """
    areas: list[tuple[int, int, int, int]] = []
    try:
        if _engine_for(xls_path) == "openpyxl":
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(filename=str(xls_path), read_only=True, data_only=True)
            ws = (
                wb[str(sheet_name)]
                if isinstance(sheet_name, str)
                else wb.worksheets[int(sheet_name)]
            )
            pa = getattr(ws, "print_area", None)
            if pa:
                try:
                    # openpyxl PrintArea has .ranges
                    for r in getattr(pa, "ranges", []):
                        areas.append((r.min_row - 1, r.max_row, r.min_col - 1, r.max_col))
                except Exception:
                    for b in _bounds_from_a1(str(pa)):
                        areas.append(b)
            # Also inspect defined names
            try:
                sheet_index = wb.sheetnames.index(ws.title)
                for dn in wb.defined_names.definedName:
                    if (
                        dn.name
                        and dn.name.lower() == "print_area"
                        and (dn.localSheetId == sheet_index or dn.localSheetId is None)
                    ):
                        for title, coord in dn.destinations:
                            if title == ws.title:
                                areas.extend(list(_bounds_from_a1(coord)))
            except Exception:
                pass
            try:
                wb.close()
            except Exception:
                pass
        else:
            import xlrd  # type: ignore

            book = xlrd.open_workbook(str(xls_path), formatting_info=True)
            idx = None
            try:
                idx = (
                    book.sheet_names().index(sheet_name)
                    if isinstance(sheet_name, str)
                    else int(sheet_name)
                )
            except Exception:
                idx = 0
            for name in book.name_map.get("print_area") or []:
                try:
                    if getattr(name, "scope", None) != idx:
                        continue
                    res = getattr(name, "result", None)
                    if res is not None:
                        vals = getattr(res, "value", [])
                        for ref3d in vals:
                            coords = getattr(ref3d, "coords", None)
                            if coords and len(coords) == 6:
                                _, _, r0, r1, c0, c1 = coords
                                areas.append((int(r0), int(r1), int(c0), int(c1)))
                            else:
                                txt = getattr(res, "text", "")
                                for b in _bounds_from_a1(txt):
                                    areas.append(b)
                except Exception:
                    continue
    except Exception:
        return []
    # Deduplicate
    uniq: list[tuple[int, int, int, int]] = []
    for a in areas:
        if a not in uniq:
            uniq.append(a)
    return uniq


def read_vertical_blocks(
    xls_path: Path,
    sheet: str | int,
    person_col: str = "A",
    regno_col: str = "B",
    blocks: list[tuple[str, str]] = [("JIS", "C:H"), ("BOILER", "I:K")],
    max_probe_rows: int = 10,
) -> pd.DataFrame:
    """Read sheets where names span multiple rows and license data sits in vertical blocks.

    - Column `person_col` contains a name only on the first row of a group; subsequent rows are blank -> forward filled.
    - Column `regno_col` similarly forward filled.
    - For each block definition (label, col_range), every row that has any non-empty value
      within that block becomes one record for that person and category.

    Returns a normalized DataFrame with columns:
      name, license_no, category, row_index, values (dict of raw block cell values),
      qualification (best-effort text), first_issue_date, issue_date, expiry_date.
    """
    # Read raw with no headers; we cannot rely on labeled headers in this layout
    _df0 = pd.read_excel(
        xls_path, sheet_name=sheet, header=None, engine=_engine_for(xls_path), dtype="object"
    )
    # Preserve original row indices for active/retired marking via print area
    mask_nonempty = ~_df0.isna().all(axis=1)
    orig_row_index = _df0.index[mask_nonempty].tolist()
    df_raw = _df0.loc[mask_nonempty].reset_index(drop=True)
    # Auto-detection mode: if caller passed special tokens or empty blocks, probe layout
    auto_person = str(person_col).strip().upper() in {"", "AUTO"}
    auto_regno = str(regno_col).strip().upper() in {"", "AUTO"}
    auto_blocks = (not blocks) or (
        len(blocks) == 1 and blocks[0][1].strip().upper() in {"AUTO", ""}
    )
    if auto_person or auto_regno or auto_blocks:
        p_idx, r_idx, detected = detect_vertical_layout_df(df_raw, max_probe_rows=max_probe_rows)
        if auto_person:
            person_col = _index_to_col_letter(p_idx)
        if auto_regno:
            regno_col = _index_to_col_letter(r_idx)
        if auto_blocks:
            blocks = [
                (lab, f"{_index_to_col_letter(a)}:{_index_to_col_letter(b - 1)}")
                for lab, (a, b) in detected
            ]
    # (already dropped empty rows above)

    p_idx = _col_letter_to_index(person_col)
    r_idx = _col_letter_to_index(regno_col)
    block_defs: list[tuple[str, tuple[int, int]]] = [
        (lab, _parse_col_range(rng)) for lab, rng in blocks
    ]

    # Helpers
    import re

    def _looks_like_regno(s: Optional[str]) -> bool:
        if s is None:
            return False
        t = str(s).strip().replace(" ", "")
        return bool(re.fullmatch(r"[0-9０-９]{1,6}(-[0-9０-９]+)?", t))

    def _looks_like_dateish(s: Optional[str]) -> bool:
        if not s:
            return False
        t = str(s)
        # quick hits: contains 年 or yyyy-mm-dd or yy.mm.dd
        return ("年" in t and ("月" in t or "日" in t)) or bool(
            re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{2}[./]\d{1,2}[./]\d{1,2}", t)
        )

    def _is_headerish_name(s: Optional[str]) -> bool:
        if not s:
            return True
        t = str(s).strip()
        if t == "":
            return True
        if any(k in t for k in ("氏名", "生年", "西暦", "証明", "溶接", "会社")):
            return True
        if t.startswith("(") and t.endswith(")"):
            return True
        if _looks_like_dateish(t):
            return True
        return False

    # Forward-fill regno anchored groups, and keep the first non-headerish Japanese-looking name in the group
    df_aux = df_raw.copy()
    for c in (p_idx, r_idx):
        if c >= df_aux.shape[1]:
            # Extend columns if necessary
            for _ in range(c - df_aux.shape[1] + 1):
                df_aux[df_aux.shape[1]] = None
    names: list[Optional[str]] = []
    regnos: list[Optional[str]] = []
    cur_name: Optional[str] = None
    cur_reg: Optional[str] = None
    seen_any_valid_reg = False
    for _, row in df_aux.iterrows():
        raw_name = row.iloc[p_idx] if p_idx < len(row) else None
        raw_reg = row.iloc[r_idx] if r_idx < len(row) else None
        # Start new person when regno cell is explicitly set and looks like a registration number
        if pd.notna(raw_reg) and str(raw_reg).strip() != "":
            candidate_reg = str(raw_reg).strip()
            if _looks_like_regno(candidate_reg):
                cur_reg = candidate_reg
                seen_any_valid_reg = True
                # choose name on new person start (if present and not headerish)
                if pd.notna(raw_name) and not _is_headerish_name(str(raw_name)):
                    cur_name = str(raw_name).strip()
            else:
                # Non-numeric header-like content in regno column — do not start a new group
                pass
        else:
            # Within a group: do not override cur_name with romanization or birthyear lines
            pass
        names.append(cur_name)
        regnos.append(cur_reg)

    # Build records per non-empty block row
    from .dates_jp import parse_jp_date

    records: list[Dict[str, Any]] = []
    for ridx, row in df_aux.reset_index(drop=True).iterrows():
        idx_i = int(ridx)  # type: ignore[arg-type,call-overload]
        name = names[idx_i]
        regno = regnos[idx_i]
        # Skip until the first valid registration appears
        if not seen_any_valid_reg or not regno or not name:
            continue
        orig_row = int(orig_row_index[idx_i]) if idx_i < len(orig_row_index) else idx_i
        for category, (c0, c1) in block_defs:
            # slice block
            vals = []
            used_cols_abs: list[int] = []
            for ci in range(c0, min(c1, df_aux.shape[1])):
                v = row.iloc[ci] if ci < len(row) else None
                vals.append(v)
                if not (pd.isna(v) or str(v).strip() == ""):
                    used_cols_abs.append(ci)
            # Skip if all empty
            if all(pd.isna(v) or str(v).strip() == "" for v in vals):
                continue

            # Build raw mapping and derive fields
            raw_map = {
                f"c{idx}": (None if pd.isna(v) else str(v).strip()) for idx, v in enumerate(vals)
            }
            # Extract dates found in any cell of this block
            date_candidates: list[pd.Timestamp] = []
            for v in raw_map.values():
                if not v:
                    continue
                try:
                    dtp = pd.to_datetime(v, errors="coerce")
                except Exception:
                    dtp = pd.NaT
                if pd.isna(dtp):
                    dtj = parse_jp_date(v)
                    dtp = pd.to_datetime(dtj, errors="coerce") if dtj else pd.NaT
                if pd.notna(dtp):
                    date_candidates.append(dtp)
            # Deduplicate and sort
            date_candidates = sorted({d.normalize() for d in date_candidates})
            first_issue: Optional[pd.Timestamp] = None
            issue: Optional[pd.Timestamp] = None
            expiry: Optional[pd.Timestamp] = None
            if len(date_candidates) >= 3:
                first_issue, issue, expiry = (
                    date_candidates[0],
                    date_candidates[-2],
                    date_candidates[-1],
                )
            elif len(date_candidates) == 2:
                first_issue, expiry = date_candidates[0], date_candidates[1]
            elif len(date_candidates) == 1:
                issue = date_candidates[0]
                # JIS heuristic: if only one date, treat as issue and set 1-year expiry
                if category.upper() == "JIS" and issue is not None:
                    try:
                        expiry = issue + pd.DateOffset(years=1)
                    except Exception:
                        expiry = None

            # Qualification guess: choose the longest non-date text token
            non_date_texts: list[str] = []
            for v in raw_map.values():
                if not v:
                    continue
                # simple guard: discard if parsing as date succeeds
                try:
                    if pd.notna(pd.to_datetime(v, errors="coerce")):
                        continue
                except Exception:
                    pass
                if any(
                    tok in v for tok in ("登録", "継続", "交付", "有効", "年月", "年", "月", "日")
                ):
                    # likely label remnants, skip
                    continue
                non_date_texts.append(v)
            qualification = max(non_date_texts, key=len) if non_date_texts else None

            # Extract issuance year from any trailing parentheses like '(23)' in the block
            import re as _re

            def _paren_year(val: Any) -> Optional[int]:
                try:
                    s = str(val)
                except Exception:
                    return None
                m = _re.search(r"\(([^)]*)\)", s)
                if not m:
                    return None
                mm = _re.search(r"(\d{2,4})", m.group(1))
                if not mm:
                    return None
                yy = mm.group(1)
                try:
                    n = int(yy)
                except Exception:
                    return None
                return (2000 + n) if len(yy) <= 2 else n

            issue_year = None
            for v in raw_map.values():
                y = _paren_year(v)
                if y is not None:
                    issue_year = y

            rec = {
                "name": name,
                "license_no": regno,
                "category": category,
                "row_index": idx_i,
                "orig_row": orig_row,
                "issue_year": issue_year,
                "block_col_start": c0,
                "block_col_end": min(c1, df_aux.shape[1]),
                "used_col_min": (min(used_cols_abs) if used_cols_abs else None),
                "used_col_max": (max(used_cols_abs) if used_cols_abs else None),
                "qualification": qualification,
                "first_issue_date": first_issue
                if (first_issue is not None and pd.notna(first_issue))
                else None,
                "issue_date": issue if (issue is not None and pd.notna(issue)) else None,
                "expiry_date": expiry if (expiry is not None and pd.notna(expiry)) else None,
                "values": raw_map,
            }
            records.append(rec)

    return pd.DataFrame.from_records(records)
